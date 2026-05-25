"""Shared Excel and table I/O helpers."""

from __future__ import annotations

from pathlib import Path
import re
from typing import Optional

import pandas as pd
from openpyxl import load_workbook


EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
MAX_EXCEL_SHEET_NAME_LENGTH = 31


def require_modern_excel_file(path: Path) -> None:
    """Raise when a file is not readable by openpyxl."""
    if path.suffix.lower() not in EXCEL_SUFFIXES:
        raise ValueError(f"Input file must be .xlsx or .xlsm: {path}")


def list_excel_sheets(input_file: Path) -> list[str]:
    """Return sheet names from a modern Excel workbook."""
    require_modern_excel_file(input_file)
    workbook = load_workbook(input_file, read_only=True)
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def sanitize_sheet_name(sheet_name: str) -> str:
    """Return a valid Excel sheet name, preserving as much text as possible."""
    cleaned = re.sub(r"[\[\]:*?/\\]", "_", sheet_name).strip()
    if not cleaned:
        cleaned = "Sheet"
    return cleaned[:MAX_EXCEL_SHEET_NAME_LENGTH]


def _trim_empty_edges(values: list[list[object]]) -> list[list[object]]:
    """Trim trailing empty rows and columns from a worksheet value grid."""
    while values and all(value is None for value in values[-1]):
        values.pop()

    if not values:
        return values

    last_column = 0
    for row in values:
        for index, value in enumerate(row, start=1):
            if value is not None:
                last_column = max(last_column, index)
    return [row[:last_column] for row in values]


def read_excel_cells_with_merged_values(
    input_file: Path,
    sheet_name: Optional[str] = None,
) -> pd.DataFrame:
    """Read a worksheet as 1-based Excel cells, filling merged-cell ranges."""
    require_modern_excel_file(input_file)
    workbook = load_workbook(input_file, data_only=True)
    try:
        worksheet = workbook.worksheets[0] if sheet_name is None else workbook[sheet_name]
        values = [[cell.value for cell in row] for row in worksheet.iter_rows()]

        for merged_range in worksheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            merged_value = values[min_row - 1][min_col - 1]
            for row_index in range(min_row - 1, max_row):
                for column_index in range(min_col - 1, max_col):
                    values[row_index][column_index] = merged_value
    finally:
        workbook.close()

    values = _trim_empty_edges(values)
    if not values:
        return pd.DataFrame()

    return pd.DataFrame(
        values,
        index=range(1, len(values) + 1),
        columns=range(1, len(values[0]) + 1),
    )


def _stringify_table(data: pd.DataFrame) -> pd.DataFrame:
    return data.map(lambda value: "" if pd.isna(value) else str(value))


def _excel_cells_to_header_table(cells: pd.DataFrame) -> pd.DataFrame:
    """Convert a 1-based worksheet-cell DataFrame into a header table."""
    if cells.empty:
        return pd.DataFrame()
    header = ["" if pd.isna(value) else str(value) for value in cells.iloc[0].tolist()]
    rows = cells.iloc[1:].copy()
    rows.columns = header
    return _stringify_table(rows.reset_index(drop=True))


def read_table(input_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
    """Read CSV or Excel as a string table with first row as headers."""
    source = Path(input_path)
    suffix = source.suffix.lower()

    if suffix == ".csv":
        return pd.read_csv(source, dtype=str, keep_default_na=False)
    if suffix in EXCEL_SUFFIXES:
        cells = read_excel_cells_with_merged_values(source, sheet_name=sheet_name)
        return _excel_cells_to_header_table(cells)
    if suffix == ".xls":
        sheet = sheet_name if sheet_name else 0
        return pd.read_excel(source, sheet_name=sheet, dtype=str, keep_default_na=False)

    raise ValueError("Input file must be a .csv, .xlsx, .xlsm, or .xls file.")


def write_table(
    data: pd.DataFrame,
    output_path: str,
    sheet_name: Optional[str] = None,
) -> Path:
    """Write results to Excel by default, or CSV when requested."""
    destination = Path(output_path)
    suffix = destination.suffix.lower()

    if suffix == ".csv":
        data.to_csv(destination, index=False)
    elif suffix in {"", ".xlsx"}:
        if suffix == "":
            destination = destination.with_suffix(".xlsx")
        if sheet_name:
            data.to_excel(destination, sheet_name=sanitize_sheet_name(sheet_name), index=False)
        else:
            data.to_excel(destination, index=False)
    else:
        raise ValueError("Output file must be .xlsx or .csv.")

    return destination


def append_or_replace_sheet(
    workbook_path: Path,
    data: pd.DataFrame,
    sheet_name: str,
) -> Path:
    """Append a DataFrame to an existing workbook, replacing that sheet if present."""
    require_modern_excel_file(workbook_path)
    sheet_name = sanitize_sheet_name(sheet_name)
    with pd.ExcelWriter(
        workbook_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace",
    ) as writer:
        data.to_excel(writer, sheet_name=sheet_name, index=False)
    return workbook_path
