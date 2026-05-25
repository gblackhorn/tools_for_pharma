"""Extract reference-gene CT QC summaries from qPCR Excel tables."""

from __future__ import annotations

import argparse
from pathlib import Path
import sys
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string, range_boundaries

from excel_utils import (
    append_or_replace_sheet,
    read_excel_cells_with_merged_values,
    sanitize_sheet_name,
    write_table,
)
from qpcr_common import (
    CONTROL_COMPOUNDS,
    MEAN_CT_COLUMN,
    REFERENCE_GENE_COLUMN,
    REFQC_SHEET_PREFIX,
    SAMPLE_SIZE_COLUMN,
    SEM_COLUMN,
    clean_text,
    is_blank,
)
from qpcr_extract import (
    DATA_START_ROW,
    HEADER_ROWS,
    TOP_GENE_ROW,
    choose_sheet_gui,
    detect_gene_names,
    find_embedded_qpcr_header_row,
    find_column,
    parse_number,
    resolved_sheet_name,
    split_side_by_side_tables,
)


INPUT_FILE = Path("qpcr_result.xlsx")
SHEET_NAME = None


class WorkbookValueResolver:
    def __init__(self, input_file: Path) -> None:
        self.value_workbook = load_workbook(input_file, data_only=True, read_only=False)
        self.formula_workbook = load_workbook(input_file, data_only=False, read_only=False)

    def close(self) -> None:
        self.value_workbook.close()
        self.formula_workbook.close()

    def resolve_cell(
        self,
        sheet_name: str,
        row: int,
        column: int,
        seen: Optional[set[tuple[str, int, int]]] = None,
    ) -> Optional[float]:
        seen = seen or set()
        key = (sheet_name, row, column)
        if key in seen:
            return None
        seen.add(key)

        value = self.value_workbook[sheet_name].cell(row, column).value
        number = parse_number(value)
        if number is not None:
            return number

        formula = self.formula_workbook[sheet_name].cell(row, column).value
        if not isinstance(formula, str) or not formula.startswith("="):
            return None
        return self.resolve_formula(sheet_name, formula[1:], seen)

    def resolve_formula(
        self,
        default_sheet: str,
        formula: str,
        seen: set[tuple[str, int, int]],
    ) -> Optional[float]:
        formula = formula.strip()
        upper_formula = formula.upper()
        if upper_formula.startswith("AVERAGE(") and formula.endswith(")"):
            values = []
            for reference in formula[8:-1].split(","):
                values.extend(self.resolve_reference(default_sheet, reference.strip(), seen))
            values = [value for value in values if value is not None]
            if not values:
                return None
            return sum(values) / len(values)

        values = self.resolve_reference(default_sheet, formula, seen)
        return values[0] if len(values) == 1 else None

    def resolve_reference(
        self,
        default_sheet: str,
        reference: str,
        seen: set[tuple[str, int, int]],
    ) -> list[Optional[float]]:
        sheet_name = default_sheet
        coordinate = reference.replace("$", "")
        if "!" in coordinate:
            sheet_part, coordinate = coordinate.split("!", 1)
            sheet_name = sheet_part.strip("'")

        if ":" in coordinate:
            min_col, min_row, max_col, max_row = range_boundaries(coordinate)
            return [
                self.resolve_cell(sheet_name, row, column, seen.copy())
                for row in range(min_row, max_row + 1)
                for column in range(min_col, max_col + 1)
            ]

        column_letters, row = coordinate_from_string(coordinate)
        column = column_index_from_string(column_letters)
        return [self.resolve_cell(sheet_name, row, column, seen.copy())]


def refqc_columns() -> list[str]:
    return [
        "Group",
        "Compound ID",
        REFERENCE_GENE_COLUMN,
        SAMPLE_SIZE_COLUMN,
        MEAN_CT_COLUMN,
        SEM_COLUMN,
    ]


def empty_refqc_summary() -> pd.DataFrame:
    return pd.DataFrame(columns=refqc_columns())


def default_refqc_sheet_name(input_file: Path, source_sheet_name: Optional[str]) -> str:
    return sanitize_sheet_name(
        f"{REFQC_SHEET_PREFIX}{resolved_sheet_name(input_file, source_sheet_name)}"
    )


def default_output_path(input_file: Path) -> Path:
    return input_file.with_name(f"{input_file.stem}_refgene_qc.xlsx")


def find_gene_mean_ct_column(data: pd.DataFrame, gene: str) -> Optional[int]:
    gene = clean_text(gene).lower()
    for column in data.columns:
        header_gene = clean_text(data.loc[TOP_GENE_ROW, column]).lower()
        subheader = clean_text(data.loc[HEADER_ROWS[-1], column]).lower()
        if header_gene == gene and "mean ct" in subheader:
            return column
    return None


def control_or_blank(group: object, compound: object) -> bool:
    return (
        is_blank(group)
        or is_blank(compound)
        or clean_text(compound).upper() in CONTROL_COMPOUNDS
    )


def extract_sample_cts_from_table(
    data: pd.DataFrame,
    table_number: int,
    resolver: WorkbookValueResolver,
    workbook_sheet_name: str,
    row_offset: int = 0,
) -> pd.DataFrame:
    genes = detect_gene_names(data)
    if len(genes) < 2:
        return pd.DataFrame()

    group_col = find_column(data, include_any=["group"])
    compound_col = find_column(data, include_any=["compound id", "compound"])
    if group_col is None or compound_col is None:
        raise ValueError(f"Could not find Group or Compound ID in table {table_number}.")

    rows = data.loc[DATA_START_ROW:].copy()
    rows[group_col] = rows[group_col].ffill()
    rows[compound_col] = rows[compound_col].ffill()

    sample_ct_rows = []
    for reference_gene in genes[1:]:
        mean_ct_col = find_gene_mean_ct_column(data, reference_gene)
        if mean_ct_col is None:
            raise ValueError(
                f"Could not find Mean CT for reference gene {reference_gene} "
                f"in table {table_number}."
            )

        for _row_index, row in rows.iterrows():
            group = row[group_col]
            compound = row[compound_col]
            mean_ct = parse_number(row[mean_ct_col])
            if mean_ct is None:
                mean_ct = resolver.resolve_cell(
                    workbook_sheet_name,
                    int(_row_index) + row_offset,
                    int(mean_ct_col),
                )
            if control_or_blank(group, compound) or mean_ct is None:
                continue
            sample_ct_rows.append(
                {
                    "Group": clean_text(group),
                    "Compound ID": clean_text(compound),
                    REFERENCE_GENE_COLUMN: reference_gene,
                    MEAN_CT_COLUMN: mean_ct,
                }
            )

    return pd.DataFrame(sample_ct_rows)


def summarize_sample_cts(sample_cts: pd.DataFrame) -> pd.DataFrame:
    if sample_cts.empty:
        return empty_refqc_summary()

    grouped = sample_cts.groupby(
        ["Group", "Compound ID", REFERENCE_GENE_COLUMN],
        sort=False,
        dropna=False,
    )[MEAN_CT_COLUMN]
    summary = grouped.agg(
        **{
            SAMPLE_SIZE_COLUMN: "count",
            MEAN_CT_COLUMN: "mean",
            SEM_COLUMN: "sem",
        }
    ).reset_index()
    summary[SEM_COLUMN] = summary[SEM_COLUMN].fillna(0)
    return summary[refqc_columns()]


def summarize_refqc_region(
    data: pd.DataFrame,
    resolver: WorkbookValueResolver,
    workbook_sheet_name: str,
    row_offset: int = 0,
) -> pd.DataFrame:
    tables = split_side_by_side_tables(data)
    full_tables = [
        (table_number, table)
        for table_number, table in enumerate(tables, start=1)
        if find_column(table, include_any=["group"]) is not None
    ]
    sample_cts = [
        extract_sample_cts_from_table(
            table,
            table_number,
            resolver,
            workbook_sheet_name,
            row_offset,
        )
        for table_number, table in full_tables
    ]
    sample_cts = [table for table in sample_cts if not table.empty]
    if not sample_cts:
        return empty_refqc_summary()
    return summarize_sample_cts(pd.concat(sample_cts, ignore_index=True))


def build_refqc_summary(
    input_file: Path,
    sheet_name: Optional[str] = SHEET_NAME,
) -> pd.DataFrame:
    workbook_sheet_name = resolved_sheet_name(input_file, sheet_name)
    resolver = WorkbookValueResolver(input_file)
    try:
        data = read_excel_cells_with_merged_values(input_file, sheet_name)
        summary = summarize_refqc_region(data, resolver, workbook_sheet_name)
        if not summary.empty:
            return summary

        header_row = find_embedded_qpcr_header_row(data)
        if header_row is None:
            return empty_refqc_summary()
        embedded_region = data.loc[header_row:].copy()
        embedded_region.index = range(1, len(embedded_region) + 1)
        return summarize_refqc_region(
            embedded_region,
            resolver,
            workbook_sheet_name,
            row_offset=header_row - 1,
        )
    finally:
        resolver.close()


def save_refqc_summary(
    summary: pd.DataFrame,
    input_file: Path,
    output_file: Optional[Path] = None,
    append_sheet: Optional[str] = None,
    output_sheet: Optional[str] = None,
) -> Path:
    if append_sheet:
        append_or_replace_sheet(input_file, summary, append_sheet)
        if output_file:
            write_table(summary, str(output_file), sheet_name=append_sheet)
            return output_file
        return input_file

    destination = output_file if output_file else default_output_path(input_file)
    return write_table(summary, str(destination), sheet_name=output_sheet)


def extract_refqc_summary(
    input_file: Path = INPUT_FILE,
    output_file: Optional[Path] = None,
    sheet_name: Optional[str] = SHEET_NAME,
    append_sheet: Optional[str] = "",
) -> pd.DataFrame:
    summary = build_refqc_summary(input_file, sheet_name)
    output_sheet = default_refqc_sheet_name(input_file, sheet_name)
    if append_sheet == "":
        append_sheet = output_sheet
    save_refqc_summary(summary, input_file, output_file, append_sheet, output_sheet)
    return summary


def run_gui() -> int:
    import tkinter as tk
    from tkinter import filedialog, messagebox, simpledialog

    root = tk.Tk()
    root.withdraw()

    try:
        input_path = filedialog.askopenfilename(
            title="Select qPCR Excel file for reference-gene QC",
            filetypes=[
                ("Excel files", "*.xlsx *.xlsm"),
                ("All files", "*.*"),
            ],
        )
        if not input_path:
            return 0

        input_file = Path(input_path)
        sheet_name = choose_sheet_gui(root, input_file)
        default_sheet_name = default_refqc_sheet_name(input_file, sheet_name)
        append_sheet = simpledialog.askstring(
            "Reference QC sheet name",
            "Sheet name:",
            initialvalue=default_sheet_name,
            parent=root,
        )
        if not append_sheet:
            return 0

        summary = extract_refqc_summary(input_file, None, sheet_name, append_sheet)
        messagebox.showinfo(
            "Done",
            f"Extracted {len(summary)} reference QC rows.\n\n"
            f"Saved to:\n{input_file}\nSheet: {append_sheet}",
        )
        return 0
    except Exception as error:
        messagebox.showerror("Reference-gene QC extraction failed", str(error))
        return 1
    finally:
        root.destroy()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Extract reference-gene Mean CT group summaries for qPCR QC."
    )
    parser.add_argument(
        "-i",
        "--input",
        type=Path,
        default=INPUT_FILE,
        help=f"Input .xlsx file. Defaults to {INPUT_FILE}.",
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help=(
            "Optional separate output .xlsx file. By default, extraction is saved "
            "as a refqc sheet appended to the input workbook."
        ),
    )
    parser.add_argument(
        "--sheet",
        default=SHEET_NAME,
        help="Excel sheet name. Defaults to the first sheet.",
    )
    parser.add_argument(
        "--append-sheet",
        nargs="?",
        const="",
        default="",
        help=(
            "Append/replace the extracted QC summary as a sheet in the input workbook. "
            "Defaults to 'refqc-[source sheet]' when no sheet name is provided."
        ),
    )
    parser.add_argument(
        "--separate-output",
        action="store_true",
        help="Save only to a separate output workbook instead of appending to input.",
    )
    parser.add_argument(
        "--gui",
        action="store_true",
        help="Choose the input workbook and sheet with dialogs.",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    if args.gui:
        return run_gui()

    append_sheet = None if args.separate_output else args.append_sheet
    output = args.output if args.output or args.separate_output else None
    try:
        summary = extract_refqc_summary(args.input, output, args.sheet, append_sheet)
    except Exception as error:
        print(f"Error: {error}", file=sys.stderr)
        return 1

    if append_sheet is not None and not output:
        appended_sheet = append_sheet if append_sheet else default_refqc_sheet_name(args.input, args.sheet)
        destination = f"{args.input} [{sanitize_sheet_name(appended_sheet)}]"
    else:
        destination = output if output else default_output_path(args.input)
    print(summary)
    print(f"Saved to: {destination}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
